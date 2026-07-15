#!/usr/bin/env python3
"""Extract the dashboard snapshot from the S&P crude-import workbook."""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl


CONTINENT_ROWS = {
    7: "Middle East",
    16: "Africa",
    38: "CIS/Europe",
    48: "Americas",
    65: "Asia-Pacific",
}
TOTAL_ROW = 81
FIRST_DATE_COLUMN = 3


def month_id(value: object) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m")
    raise ValueError(f"Unexpected month header: {value!r}")


def clean_value(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 3)
    raise ValueError(f"Unexpected data value: {value!r}")


def extract(source: Path) -> dict:
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=False)
    sheet = next(
        (candidate for candidate in workbook.worksheets if candidate.title.startswith("Crude imports")),
        None,
    )
    if sheet is None:
        raise ValueError("Could not find the 'Crude imports — Volume' worksheet")

    dates: list[str] = []
    date_columns: list[int] = []
    column = FIRST_DATE_COLUMN
    while sheet.cell(6, column).value is not None:
        dates.append(month_id(sheet.cell(6, column).value))
        date_columns.append(column)
        column += 1

    if not dates:
        raise ValueError("No month headers found from C6")

    series: list[dict] = []
    current_continent: str | None = None
    for row in range(7, TOTAL_ROW + 1):
        raw_name = sheet.cell(row, 2).value
        if not raw_name:
            continue
        name = str(raw_name).strip()
        if row == TOTAL_ROW:
            level = "total"
            continent = None
        elif row in CONTINENT_ROWS:
            level = "continent"
            current_continent = CONTINENT_ROWS[row]
            continent = current_continent
        else:
            level = "country"
            continent = current_continent

        series.append(
            {
                "id": "total" if level == "total" else name.lower().replace(" ", "-").replace("/", "-"),
                "name": name,
                "level": level,
                "continent": continent,
                "values": [clean_value(sheet.cell(row, col).value) for col in date_columns],
            }
        )

    return {
        "metadata": {
            "sourceFile": source.name,
            "sheet": sheet.title,
            "unit": "kbd",
            "sourceRange": "B6 through row 81; requested market rows 7–80 plus Total crude imports at row 81",
            "compiled": str(sheet.cell(84, 2).value or "").strip(),
            "sources": str(sheet.cell(85, 2).value or "").strip(),
            "firstMonth": dates[0],
            "latestMonth": dates[-1],
        },
        "dates": dates,
        "series": series,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    payload = extract(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(
        f"Wrote {len(payload['series'])} series × {len(payload['dates'])} months "
        f"to {args.output}"
    )


if __name__ == "__main__":
    main()
